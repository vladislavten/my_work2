import openpyxl
import os


def main():
    # Для начала постараемся найти ваш драгоценный файлик data.xlsx
    # и загрузим его, чтобы стырить имена и номера.
    try:
        wb_data = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        print("Слышь, 'data.xlsx' потерялся. Проверь, куда ты его запихнул.")
        return

    ws_data = wb_data.active  # Берём первый лист, попроще будет

    # На всякий случай проверяем, вдруг файла 'shablon.xlsx' тоже нет
    try:
        openpyxl.load_workbook('shablon.xlsx')
    except FileNotFoundError:
        print("А 'shablon.xlsx' куда дел? Без шаблона всё грустно.")
        return

    # Убедимся, что папка results существует
    if not os.path.exists('results'):
        os.makedirs('results')

    # Проходимся по данным. Предполагается, что у нас в A - имя, в B - номер,
    # и, по-хорошему, стоит начать с 2 строки, если в первой шапка.
    for row in ws_data.iter_rows(min_row=1, values_only=True):
        # row будет кортежем вида (имя, номер, ...)
        name, client_id, project_name = row[0], row[1], row[2]

        # Открываем ваш святой шаблон
        wb_template = openpyxl.load_workbook('shablon.xlsx')
        ws_template = wb_template.active

        # Пихаем имя в ячейку B2 и номер в C
        ws_template['B10'] = name
        ws_template['C10'] = client_id

        # Формируем название итогового файла. Попробуй не написать ничего странного в ячейке с именем
        output_filename = os.path.join('results', f"{project_name}_{name}.xlsx")

        # Сохраняем новый файлик
        wb_template.save(output_filename)
        print(f"Файл '{output_filename}' сохранён.")


if __name__ == "__main__":
    main()
